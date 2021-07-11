# import openpyxl as xl
from xlrd import open_workbook
from openpyxl import Workbook
import os


def read_rtc_data(source_file):
    book = open_workbook(source_file)
    sheet1 = book.sheet_by_name("Sheet1")
    sheet2 = book.sheet_by_name("Sheet2")

    rtc_date = sheet1.cell(3, 8).value
    rtc_sr_no = sheet1.cell(6, 2).value
    rtc_type = sheet1.cell(6, 6).value
    rtc_cl_curr = sheet1.cell(43, 3).value
    rtc_cl_time = sheet1.cell(45, 3).value
    rtc_op_curr = sheet2.cell(32, 3).value
    rtc_op_time = sheet2.cell(34, 3).value
    rtc_c_force1 = sheet2.cell(39, 5).value
    rtc_c_force2 = sheet2.cell(39, 6).value
    rtc_c_force3 = sheet2.cell(39, 7).value
    rtc_o_force1 = sheet2.cell(41, 5).value
    rtc_o_force2 = sheet2.cell(41, 6).value
    rtc_o_force3 = sheet2.cell(41, 7).value
    ws.append([rtc_date, rtc_sr_no, rtc_type, rtc_cl_curr,
                   rtc_cl_time, rtc_op_curr, rtc_op_time, rtc_c_force1, rtc_c_force2,
                   rtc_c_force3, rtc_o_force1, rtc_o_force2, rtc_o_force3])


wb = Workbook()
# grab the active worksheet
ws = wb.active
ws.append(['Date', 'Srqq. No.', 'Type', 'Close peak current', 'time cl_pk', 'Open peak current', 'time op_pk',
           'h_force_cl1', 'h_force_cl2', 'h_force_cl3', 'h_force_op1', 'h_force_op2', 'h_force_op3'])


os.chdir('/home/boss/PycharmProjects/RTC/data')
s_path = os.getcwd()
print(s_path)
list_of_item = os.listdir()


count = 0
for element in list_of_item:
    f_name = list_of_item[count]
    print(f"reading data from {f_name}")
    read_rtc_data(f_name)
    # write_rtc_data()
    count += 1


wb.save("/home/boss/PycharmProjects/RTC/result.xlsx")
